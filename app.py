import os
import json
import logging
from datetime import datetime
from flask import Flask, jsonify, request
import openpyxl
import webbrowser
from threading import Timer
import requests
from bs4 import BeautifulSoup

app = Flask(__name__)
# Suppress the default flask logging to keep console clean
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

EXCEL_FILE = 'Auto Data submission.xlsx'

HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
  <style>
    body { font-family: sans-serif; padding: 15px; background: #f4f4f9; }
    .card { background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); margin-bottom: 10px; }
    select, button, input[type="text"] { width: 100%; padding: 10px; margin-top: 10px; cursor: pointer; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box;}
    textarea { width: 100%; height: 200px; margin-top: 10px; font-family: monospace; font-size: 11px; box-sizing: border-box; padding: 10px;}
    .label { font-weight: bold; font-size: 14px; color: #333; }
    .btn-copy { background: #4CAF50; color: white; border: none; border-radius: 4px; padding: 10px; margin-top: 10px; cursor: pointer; font-weight: bold;}
    .btn-copy:hover { background: #45a049; }
    .loader { display: none; color: #666; font-size: 12px; margin-top: 5px; }
  </style>
</head>
<body>

  <div class="card">
    <div class="label">Step 1: Select Category</div>
    <select id="category" onchange="toggleView()">
      <option value="BOQ">BOQ (Column Based)</option>
      <option value="JV Partner">JV Partner (Async Typing)</option>
      <option value="Scrape">Scrape Payment Data link</option>
      <option value="ExtractFiles">Extract File Name from Links</option>
      <option value="Payment">Payment (Async Typing)</option>
    </select>
  </div>

  <div id="boq-section" class="card">
    <div class="label">Step 2: Select Prefix</div>
    <select id="prefixSelect"><option>Loading...</option></select>
    <button onclick="generateBOQ()">Generate BOQ Code</button>
  </div>

  <div id="scrape-section" class="card" style="display:none;">
    <div class="label">Step 1: Enter Extraction Link</div>
    <input type="text" id="extractionLink" placeholder="Paste the ViewBOQForms.jsp URL here..." style="width: 100%; padding: 9px; margin-top: 10px; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box;">
    <button onclick="scrapeData()" style="margin-top: 10px;">Scrape & Extract All</button>
    <p style="font-size: 11px; color: #666; margin-top: 5px;">This will login, scrape table data, and automatically extract filenames.</p>
    <p style="font-size: 11px; color: #d32f2f; margin-top: 5px; background: #ffebee; padding: 5px; border-radius: 4px;">
      <b>Tip:</b> If right-click is disabled, use <code>Ctrl+C</code> or paste this in Console: <br>
      <code>document.oncontextmenu=null;document.onselectstart=null;</code>
    </p>
  </div>

  <div id="async-section" class="card" style="display:none;">
    <div id="payment-link-row" style="margin-bottom: 20px;">
       <div class="label">Step 1: Enter Payment Link</div>
       <input type="text" id="paymentLink" placeholder="Paste payment page URL here..." style="width: 100%; padding: 9px; margin-top: 10px; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box;">
    </div>
    <div style="margin-bottom: 20px;">
       <div class="label">Step 2: Enter ID Prefix (Optional)</div>
       <input type="text" id="idPrefix" placeholder="e.g. 9807325" style="width: 100%; padding: 9px; margin-top: 10px; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box;">
       <p style="font-size: 11px; color: #666; margin-top: 5px;">Required if the page IDs look like 'bttnXXXXXXX_...'.</p>
    </div>
    <button onclick="submitPayment()" style="margin-top: 10px; background: #2196F3; color: white;">Start Automated Submission</button>
    <p style="font-size: 11px; color: #666; margin-top: 5px;">This will open Chrome browser. Log in first, then it will fill the form automatically.</p>
    <p style="font-size: 11px; color: #d32f2f; margin-top: 5px; background: #ffebee; padding: 5px; border-radius: 4px;">
      <b>Tip:</b> If right-click is disabled on eProcure, use <code>Ctrl+C</code> or paste this in Console (F12): <br>
      <code>document.oncontextmenu=null;document.onselectstart=null;</code>
    </p>
    <div id="async-code-row">
       <div class="label">Step 2: Confirm Generation</div>
       <button onclick="generateAsync()">Generate Async Code (Manual Console)</button>
    </div>
  </div>

  <div id="scrape-section" class="card" style="display:none;">
    <div class="label">Step 2: Paste eProcure Page HTML Source (Ctrl+U, Ctrl+A, Ctrl+C)</div>
    <textarea id="scrapeHtml" placeholder="Paste full HTML here..." style="width: 100%; height: 100px; padding: 10px; margin-top: 10px; cursor: text; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box;"></textarea>
    <button onclick="scrapeData()">Scrape Data</button>
  </div>

  <div id="extract-section" class="card" style="display:none;">
    <div class="label">Step 2: Run Extraction</div>
    <p style="font-size: 13px; color: #555;">This will open a Chrome browser window. <b>Make sure you are logged into eProcure in that browser window</b> before clicking Start.</p>
    <p style="font-size: 13px; color: #e67e00; background: #fff3cd; padding: 5px; border-radius: 4px;"><b>Do not close the Chrome window</b> while it is running. Close your Excel file first to avoid save conflicts.</p>
    <button onclick="extractFiles()">Start Extraction</button>
  </div>

  <div class="card">
    <div class="label">Generated Code / Status:</div>
    <div id="loader" class="loader">Processing...</div>
    <textarea id="output" readonly placeholder="Code will appear here..."></textarea>
    <button class="btn-copy" onclick="copyCode(this)">Copy to Clipboard</button>
  </div>

  <script>
    // Initialize prefixes for BOQ
    fetch('/getPrefixes')
      .then(r => r.json())
      .then(prefixes => {
        const select = document.getElementById('prefixSelect');
        if (prefixes.error) {
           select.innerHTML = '<option>' + prefixes.error + '</option>';
        } else {
           select.innerHTML = prefixes.map(p => `<option value="${p}">${p}</option>`).join('');
        }
      })
      .catch(e => {
        console.error(e);
        document.getElementById('prefixSelect').innerHTML = '<option>Error loading</option>';
      });

    function toggleView() {
      const cat = document.getElementById('category').value;
      document.getElementById('boq-section').style.display = cat === 'BOQ' ? 'block' : 'none';
      document.getElementById('async-section').style.display = (cat === 'JV Partner' || cat === 'Payment') ? 'block' : 'none';
      document.getElementById('scrape-section').style.display = cat === 'Scrape' ? 'block' : 'none';
      document.getElementById('extract-section').style.display = cat === 'ExtractFiles' ? 'block' : 'none';
      
      // Toggle Payment Link input
      document.getElementById('payment-link-row').style.display = cat === 'Payment' ? 'block' : 'none';
      document.getElementById('async-code-row').style.display = cat === 'JV Partner' ? 'block' : 'none';
    }

    function showLoader(show) {
      document.getElementById('loader').style.display = show ? 'block' : 'none';
      document.getElementById('output').style.opacity = show ? '0.5' : '1';
    }

    function generateBOQ() {
      const p = document.getElementById('prefixSelect').value;
      showLoader(true);
      fetch('/generateBOQCode?prefix=' + encodeURIComponent(p))
        .then(r => r.text())
        .then(code => {
          document.getElementById('output').value = code;
          showLoader(false);
        }).catch(() => showLoader(false));
    }

    function generateAsync() {
      const cat = document.getElementById('category').value;
      showLoader(true);
      fetch('/generateHandsOffJS?sheetName=' + encodeURIComponent(cat))
        .then(r => r.text())
        .then(code => {
          document.getElementById('output').value = code;
          showLoader(false);
        }).catch(() => showLoader(false));
    }

    async function submitPayment() {
      const url = document.getElementById('paymentLink').value;
      const prefix = document.getElementById('idPrefix').value;
      if (!url) {
          alert('Please paste the payment link first');
          return;
      }
      showLoader(true);
      const out = document.getElementById('output');
      out.value = "Opening Chrome browser...\\nPlease log in to eProcure in the window that opens, then submission will start automatically.\\n";

      try {
        const response = await fetch('/submitPayment', {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({url: url, prefix: prefix})
        });
        const reader = response.body.getReader();
        const decoder = new TextDecoder();
        let buffer = '';

        while (true) {
          const {done, value} = await reader.read();
          if (done) break;
          buffer += decoder.decode(value, {stream: true});
          const parts = buffer.split('\\n\\n');
          buffer = parts.pop();
          for (const part of parts) {
            const line = part.replace(/^data: /, '');
            if (line) out.value += line + '\\n';
          }
          out.scrollTop = out.scrollHeight;
        }
      } catch (e) {
        out.value += '\\n❌ Connection error: ' + e;
      }
      showLoader(false);
    }
    async function scrapeData() {
      const link = document.getElementById('extractionLink').value;
      if (!link) { alert('Please enter the extraction link.'); return; }
      const out = document.getElementById('output');
      out.value = "Launching browser for automated scrape & extract...\\nPlease log in first...\\n";
      showLoader(true);
      try {
        const response = await fetch('/scrapeData', {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({link})
        });
        const reader = response.body.getReader();
        const decoder = new TextDecoder();
        let buffer = '';
        while (true) {
          const {done, value} = await reader.read();
          if (done) break;
          buffer += decoder.decode(value, {stream: true});
          const parts = buffer.split('\\n\\n');
          buffer = parts.pop();
          for (const part of parts) {
            const line = part.replace(/^data: /, '');
            if (line) out.value += line + '\\n';
          }
          out.scrollTop = out.scrollHeight;
        }
      } catch (e) {
        out.value += "\\n❌ Connection error: " + e;
      }
      showLoader(false);
    }

    async function extractFiles() {
      showLoader(true);
      const out = document.getElementById('output');
      out.value = "Opening Chrome browser...\\nPlease log in to eProcure in the window that opens, then scraping will start automatically.\\n";

      try {
        const response = await fetch('/extractFiles', {method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({})});
        const reader = response.body.getReader();
        const decoder = new TextDecoder();
        let buffer = '';

        while (true) {
          const {done, value} = await reader.read();
          if (done) break;
          buffer += decoder.decode(value, {stream: true});
          const parts = buffer.split('\\n\\n');
          buffer = parts.pop(); // keep incomplete chunk
          for (const part of parts) {
            const line = part.replace(/^data: /, '');
            if (line) out.value += line + '\\n';
          }
          out.scrollTop = out.scrollHeight;
        }
      } catch (e) {
        out.value += '\\n❌ Connection error: ' + e;
      }
      showLoader(false);
    }

    function copyCode(button) {
      const copyText = document.getElementById("output");
      copyText.select();
      document.execCommand("copy");

      const originalText = button.innerText;
      button.innerText = "Code copied!";

      setTimeout(() => {
        button.innerText = originalText;
      }, 2500); // 2.5 seconds
    }

    // Initialize correct section visibility on page load
    toggleView();
  </script>
</body>
</html>
"""

def get_workbook():
    if not os.path.exists(EXCEL_FILE):
        return None
    try:
        # data_only=True ensures we read calculated values instead of formulas
        return openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return None

def format_date(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val).strip()

def buildAsyncJS(prefix, rows):
    rows_json = json.dumps(rows, default=str)
    
    code = f"(async function () {{\n"
    code += f"    window.confirm = () => true;\n"
    code += f"    const prefix = '{prefix}';\n"
    code += f"    const rows = {rows_json};\n"
    code += f"    const wait = ms => new Promise(res => setTimeout(res, ms));\n"
    code += f"    const typeValue = async (el, value) => {{\n"
    code += f"      if (!el) return;\n"
    code += f"      el.focus(); el.value = '';\n"
    code += f"      for (let char of String(value)) {{\n"
    code += f"        el.value += char;\n"
    code += f"        el.dispatchEvent(new Event('input', {{ bubbles: true }}));\n"
    code += f"        await wait(15);\n"
    code += f"      }}\n"
    code += f"      el.dispatchEvent(new Event('change', {{ bubbles: true }}));\n"
    code += f"      el.dispatchEvent(new Event('blur', {{ bubbles: true }}));\n"
    code += f"    }};\n"
    code += f"    for (let i = 0; i < rows.length; i++) {{\n"
    code += f"      const rowNum = i + 1; const r = rows[i];\n"
    code += f"      const fill = async (colIdx, val, isDate = false) => {{\n"
    code += f"        const targetId = 'row' + prefix + '_' + rowNum + '_' + colIdx;\n"
    code += f"        const el = document.getElementById(targetId);\n"
    code += f"        if (el) {{\n"
    code += f"          if (isDate) {{ el.value = val; el.dispatchEvent(new Event('change', {{ bubbles: true }})); }}\n"
    code += f"          else {{ await typeValue(el, val); }}\n"
    code += f"        }}\n"
    code += f"      }};\n"
    code += f"      await fill(1, r.slNo); await fill(2, r.tid); await fill(3, r.amount);\n"
    code += f"      await fill(4, r.dateNoa, true); await fill(5, r.dateComp, true);\n"
    code += f"      await fill(6, r.peName); await fill(7, r.payment);\n"
    code += f"      if (i < rows.length - 1) {{\n"
    code += f"        const btn = document.getElementById('bttn' + prefix);\n"
    code += f"        if (btn) {{ btn.click(); await wait(1200); }}\n"
    code += f"      }}\n"
    code += f"    }}\n"
    code += f"    console.log('✅ Success');\n"
    code += f"  }})();"
    return code

@app.route('/')
def index():
    return HTML_TEMPLATE

@app.route('/getPrefixes')
def get_prefixes():
    wb = get_workbook()
    if not wb:
        return jsonify({"error": "Auto Data submission.xlsx not found"})
    if 'BOQ' not in wb.sheetnames:
        return jsonify({"error": "Sheet BOQ not found in the excel file"})
        
    sheet = wb['BOQ']
    data = list(sheet.iter_rows(values_only=True))
    if not data or not data[0]:
        return jsonify([])
        
    # The original script does data[0].slice(1), so we skip the first column
    prefixes = [str(x) for x in data[0][1:] if x is not None and str(x).strip() != '']
    return jsonify(prefixes)

@app.route('/generateBOQCode')
def generate_boq_code():
    prefix = request.args.get('prefix', '')
    if not prefix:
        return "❌ Missing prefix"
        
    wb = get_workbook()
    if not wb:
        return "❌ Excel file not found"
    if 'BOQ' not in wb.sheetnames:
        return "❌ Sheet BOQ not found"
        
    sheet = wb['BOQ']
    data = list(sheet.iter_rows(values_only=True))
    if not data:
        return "❌ No data in BOQ sheet"
        
    headerRow = data[0]
    col_index = -1
    for i, cell in enumerate(headerRow):
        if cell is not None and str(cell).strip() == str(prefix).strip():
            col_index = i
            break
            
    if col_index == -1:
        return f"❌ Prefix '{prefix}' not found in BOQ headers"
        
    jsCode = f"// BOQ Injector for Prefix: {prefix}\n"
    for r in range(1, len(data)):
        row_data = tuple(data[r])
        value = row_data[col_index] if col_index < len(row_data) else None
        if value is None:
            value = ""
        
        # Format the value for javascript string injection
        val_str = str(value).replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n').replace('\r', '')
        
        inputId = f"row{prefix}_{r}_{7}"
        jsCode += f'let el{r} = document.getElementById("{inputId}"); if(el{r}){{ el{r}.value="{val_str}"; el{r}.dispatchEvent(new Event(\'blur\')); }}\n'
        
    return jsCode

@app.route('/generateHandsOffJS')
def generate_hands_off_js():
    sheetName = request.args.get('sheetName', '')
    if not sheetName:
        return "❌ Missing sheetName"
        
    wb = get_workbook()
    if not wb:
        return "❌ Excel file not found"
    if sheetName not in wb.sheetnames:
        return f"❌ Sheet '{sheetName}' not found"
        
    sheet = wb[sheetName]
    
    # In original Google Apps Script:
    # const prefix = sheet.getRange("A2").getValue();
    prefix_val = sheet['A2'].value
    prefix = str(prefix_val).strip() if prefix_val is not None else ''
    
    # Original logic: getRange(2, 2, lastRow - 1, 8).getValues()
    # Which means we iterate over columns B through I (2 to 9), starting at row 2
    rows = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=9, values_only=True):
        if row[0] is not None and str(row[0]).strip() != '':
            slNo = row[0]
            tid = row[1] if len(row) > 1 and row[1] is not None else ''
            amount = row[2] if len(row) > 2 and row[2] is not None else ''
            dateNoa = format_date(row[3]) if len(row) > 3 else ''
            dateComp = format_date(row[4]) if len(row) > 4 else ''
            peName = row[5] if len(row) > 5 and row[5] is not None else ''
            payment = row[6] if len(row) > 6 and row[6] is not None else ''
            
            rows.append({
                'slNo': slNo,
                'tid': tid,
                'amount': amount,
                'dateNoa': dateNoa,
                'dateComp': dateComp,
                'peName': peName,
                'payment': payment
            })
            
    return buildAsyncJS(prefix, rows)

@app.route('/scrapeData', methods=['POST'])
def scrape_data():
    req_data = request.get_json()
    target_url = req_data.get('link')
    
    if not target_url:
        return jsonify({"error": "No link provided"}), 400

    from flask import Response, stream_with_context

    def generate():
        yield "data: 🚀 Starting Automated Scrape & Extract Chain...\n\n"
        
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager
        import time
        import re

        opts = Options()
        opts.add_argument("--start-maximized")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

        try:
            # 1. Login
            driver.get("https://www.eprocure.gov.bd/Index.jsp")
            yield "data: ⏳ Waiting for login at Index.jsp (up to 3 minutes)...\n\n"
            
            deadline = time.time() + 180
            logged_in = False
            while time.time() < deadline:
                cur = driver.current_url
                if 'Index.jsp' not in cur and 'Login' not in cur and 'eprocure.gov.bd' in cur:
                    logged_in = True
                    break
                time.sleep(1)

            if not logged_in:
                yield "data: ❌ Login timed out.\n\n"
                return

            yield "data: ✅ Login detected! Navigating to extraction link...\n\n"
            # Simulated click to bypass security
            driver.execute_script(f"var a=document.createElement('a');a.href='{target_url}';a.style.display='none';document.body.appendChild(a);a.click();")
            
            # Wait for either mtable1 or tableList_1
            try:
                WebDriverWait(driver, 10).until(
                    lambda d: d.find_elements(By.ID, 'mtable1') or d.find_elements(By.CLASS_NAME, 'tableList_1')
                )
            except:
                yield "data: ❌ Timeout waiting for table structure (mtable1 or tableList_1).\n\n"
                return

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            
            # Try finding the nested data table (for BOQ pages)
            table1 = soup.find('table', id='mtable1')
            actual_data_table = None
            if table1:
                trs_all = table1.find_all('tr', recursive=False)
                if not trs_all:
                    tbody = table1.find('tbody', recursive=False)
                    if tbody: trs_all = tbody.find_all('tr', recursive=False)
                if len(trs_all) >= 2:
                    actual_data_table = trs_all[-2].find('table')
            
            if not actual_data_table:
                actual_data_table = soup.find('table', class_='tableList_1')

            if not actual_data_table:
                yield "data: ❌ Could not find a valid data table on the page.\n\n"
                return

            scraped_rows = []
            rows_raw = actual_data_table.find_all('tr')
            start_idx = 2 if table1 else 0
            
            for tr in rows_raw[start_idx:]:
                tds = tr.find_all('td')
                if tds:
                    cols = [td.get_text(strip=True) for td in tds]
                    if len(cols) >= 8:
                        # Find the "View Mapped Documents" link
                        map_link_url = ""
                        for a in tr.find_all('a'):
                            if "View Mapped Documents" in a.get_text():
                                # Extract from onclick or href
                                href = a.get('href', '')
                                if href.startswith('../../'):
                                    map_link_url = "https://www.eprocure.gov.bd/" + href.replace('../../', '')
                                break
                        
                        scraped_rows.append({
                            'slNo': cols[0],
                            'tid': cols[1],
                            'amount': cols[2],
                            'dateNoa': cols[3],    # Index 3 -> 4th td -> Column D
                            'dateComp': cols[4],   # Index 4 -> 5th td -> Column E
                            'peName': cols[5],     # Index 5 -> 6th td -> Column F
                            'payment': cols[6],    # Index 6 -> 7th td -> Column G
                            'mapUrl': map_link_url
                        })

            if not scraped_rows:
                yield "data: ❌ Scraped 0 rows.\n\n"
                return

            yield f"data: ✅ Scraped {len(scraped_rows)} rows. Extracting file names...\n\n"
            
            wb = get_workbook()
            if not wb:
                yield "data: ❌ Excel file not found.\n\n"
                return
            if 'Payment' not in wb.sheetnames:
                wb.create_sheet('Payment')
            sheet = wb['Payment']
            
            # Header Row
            sheet.delete_rows(1, sheet.max_row)
            headers = ["Sl No", "TID or Ref", "Amount", "NOA Date", "Comp Date", "PE Name", "Payment", "Remaining", "Upload File"]
            for c, h in enumerate(headers, 1):
                sheet.cell(row=1, column=c).value = h
            
            main_h = driver.current_window_handle
            
            for i, r in enumerate(scraped_rows, 1):
                yield f"data: [{i}/{len(scraped_rows)}] Extracting row {r['slNo']}...\n\n"
                
                filename = ""
                if r['mapUrl']:
                    try:
                        driver.execute_script(f"window.open('{r['mapUrl']}', '_blank');")
                        time.sleep(2)
                        for h in driver.window_handles:
                            if h != main_h:
                                driver.switch_to.window(h)
                                break
                        
                        WebDriverWait(driver, 7).until(EC.presence_of_element_located((By.CLASS_NAME, 'tableList_1')))
                        inner_soup = BeautifulSoup(driver.page_source, 'html.parser')
                        inner_tab = inner_soup.find('table', class_='tableList_1')
                        if inner_tab:
                            trs_inner = [tr for tr in inner_tab.find_all('tr') if tr.find('td')]
                            if trs_inner:
                                # Last tr, 2nd td
                                last_tds = trs_inner[-1].find_all('td')
                                if len(last_tds) >= 2:
                                    filename = last_tds[1].get_text(strip=True)
                        
                        driver.close()
                        driver.switch_to.window(main_h)
                    except Exception as e:
                        yield f"data: ⚠️ Error extracting filename: {str(e)[:50]}\n\n"
                        # Ensure we switch back even on error
                        if len(driver.window_handles) > 1:
                            driver.close()
                        driver.switch_to.window(main_h)

                # Save Row to Excel
                row_idx = i + 1
                sheet.cell(row=row_idx, column=1).value = r['slNo']
                sheet.cell(row=row_idx, column=2).value = r['tid']
                sheet.cell(row=row_idx, column=3).value = r['amount']
                sheet.cell(row=row_idx, column=4).value = r['dateNoa']
                sheet.cell(row=row_idx, column=5).value = r['dateComp']
                sheet.cell(row=row_idx, column=6).value = r['peName']
                sheet.cell(row=row_idx, column=7).value = r['payment']
                
                # Calculate Remaining (8 = 3-7)
                try:
                    amt = float(re.sub(r'[^\d.]', '', str(r['amount'])))
                    pay = float(re.sub(r'[^\d.]', '', str(r['payment'])))
                    sheet.cell(row=row_idx, column=8).value = amt - pay
                except:
                    sheet.cell(row=row_idx, column=8).value = "Error"
                
                sheet.cell(row=row_idx, column=9).value = filename
                
                yield f"data: ✅ Row {i}: Saved ({filename if filename else 'No file'})\n\n"

            wb.save(EXCEL_FILE)
            yield f"data: \n🎉 Full Scrape Complete! Saved {len(scraped_rows)} rows to Columns A-I.\n\n"
            yield "data: ℹ️ Browser remains open.\n\n"

        except Exception as global_e:
            yield f"data: ❌ Chain failed: {global_e}\n\n"
        finally:
            pass

    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/submitPayment', methods=['POST'])
def submit_payment():
    from flask import Response, stream_with_context
    data = request.json
    target_url = data.get('url', '')
    prefix = data.get('prefix', '')

    def generate():
        wb = get_workbook()
        if not wb or 'Payment' not in wb.sheetnames:
            yield "data: ❌ 'Payment' sheet not found in Excel.\n\n"
            return
        
        sheet = wb['Payment']
        # Use prefix passed from UI (line 647)
        # We no longer read it from A2 as A2 is Sl No now.
        # I'll default to empty prefix if not sure.
        
        rows = []
        # We start from row 2 as row 1 is header
        for r_idx in range(2, sheet.max_row + 1):
            sl = sheet.cell(row=r_idx, column=1).value
            if sl is not None and str(sl).strip() != '':
                rows.append({
                    'slNo': sl,
                    'tid': sheet.cell(row=r_idx, column=2).value or '',
                    'amount': sheet.cell(row=r_idx, column=3).value or '',
                    'dateNoa': format_date(sheet.cell(row=r_idx, column=4).value),
                    'dateComp': format_date(sheet.cell(row=r_idx, column=5).value),
                    'peName': sheet.cell(row=r_idx, column=6).value or '',
                    'payment': sheet.cell(row=r_idx, column=7).value or '',
                    'remaining': sheet.cell(row=r_idx, column=8).value or '',
                    'fileName': str(sheet.cell(row=r_idx, column=9).value or '').strip()
                })

        if not rows:
            yield "data: ❌ No data rows found in Payment sheet.\n\n"
            return

        yield f"data: Found {len(rows)} rows to submit. Launching Chrome (Keep Browser Open = Enabled)...\n\n"

        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from webdriver_manager.chrome import ChromeDriverManager
            import time

            opts = Options()
            opts.add_argument("--start-maximized")
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
        except Exception as e:
            yield f"data: ❌ Failed to launch Chrome: {e}\n\n"
            return

        try:
            # Login first as requested
            driver.get("https://www.eprocure.gov.bd/Index.jsp")
            yield "data: ⏳ Waiting for you to log in (up to 3 minutes)...\n\n"

            deadline = time.time() + 180
            logged_in = False
            while time.time() < deadline:
                cur = driver.current_url
                if 'Index.jsp' not in cur and 'Login' not in cur and 'eprocure.gov.bd' in cur:
                    logged_in = True
                    break
                time.sleep(1)

            if not logged_in:
                yield "data: ❌ Login timed out (3 minutes). Please restart and log in faster.\n\n"
                return

            yield "data: ✅ Login detected! Navigating to payment link via simulated click...\n\n"
            # Using simulated click to preserve Referer/Session
            driver.execute_script(f"var a=document.createElement('a');a.href='{target_url}';a.style.display='none';document.body.appendChild(a);a.click();")
            time.sleep(3) # Wait for page load
            # Suppress confirmation alerts
            driver.execute_script("window.confirm = function() { return true; }; window.alert = function() { return true; };")
            time.sleep(1)
            
            yield f"data: 🚀 Proceeding to fill and map {len(rows)} rows...\n\n"

            for i, r in enumerate(rows):
                rowNum = i + 1
                yield f"data: [{rowNum}/{len(rows)}] Filling row fields...\n\n"
                
                def fill_field(col_idx, val):
                    target_id = f"row{prefix}_{rowNum}_{col_idx}"
                    script = f"""
                        var el = document.getElementById("{target_id}");
                        if (el) {{
                            el.value = "{str(val).replace('"', '\\"').replace('\\n', '\\\\n')}";
                            el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                            el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            el.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                        }}
                    """
                    driver.execute_script(script)

                fill_field(1, r['slNo'])
                fill_field(2, r['tid'])
                fill_field(3, r['amount'])
                fill_field(4, r['dateNoa'])
                fill_field(5, r['dateComp'])
                fill_field(6, r['peName'])
                fill_field(7, r['payment'])
                
                # --- Document Mapping Logic ---
                file_name = r['fileName']
                if file_name:
                    yield f"data: [{rowNum}/{len(rows)}] Mapping document: {file_name}...\n\n"
                    link_id = f"link{prefix}_{rowNum}"
                    try:
                        main_window = driver.current_window_handle
                        # Find and click the link
                        map_link = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, link_id)))
                        map_link.click()
                        time.sleep(2)
                        
                        # Wait for popup and switch to it
                        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
                        for handle in driver.window_handles:
                            if handle != main_window:
                                driver.switch_to.window(handle)
                                break
                        
                        # Click Search button
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "search_list"))).click()
                        
                        # Fill search field
                        search_inp = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input.vdata")))
                        search_inp.clear()
                        search_inp.send_keys(file_name)
                        
                        # Click Find
                        find_btn = driver.find_element(By.CSS_SELECTOR, ".ui-search")
                        find_btn.click()
                        time.sleep(1.5)
                        
                        # Close Search Section
                        driver.find_element(By.CSS_SELECTOR, ".ui-closer").click()
                        time.sleep(1)
                        
                        # Check the box
                        chk = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "chkComDocId")))
                        if not chk.is_selected():
                            chk.click()
                        
                        # Click Map
                        driver.find_element(By.ID, "btnMap").click()
                        
                        # Wait for success message
                        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, "successMsg")))
                        
                        # Close and switch back
                        driver.close()
                        driver.switch_to.window(main_window)
                        yield f"data: [{rowNum}/{len(rows)}] Success: Document mapped.\n\n"
                    except Exception as map_err:
                        yield f"data: ⚠️  [{rowNum}/{len(rows)}] Map failed: {str(map_err)}\n\n"
                        # Recover to main window
                        if len(driver.window_handles) > 1:
                            driver.close()
                        driver.switch_to.window(main_window)

                if i < len(rows) - 1:
                    btn_id = f"bttn{prefix}"
                    try:
                        btn = driver.find_element(By.ID, btn_id)
                        btn.click()
                        # Longer wait to ensure new row is fully rendered
                        time.sleep(2.5)
                        # Re-suppress alerts in case of page reload/navigation
                        driver.execute_script("window.confirm = function() { return true; }; window.alert = function() { return true; };")
                    except Exception as btn_err:
                        yield f"data: ⚠️ Could not find or click 'Add More' button (ID: {btn_id}). Stopping. Error: {str(btn_err)}\n\n"
                        break
            
            yield f"data: 🎉 Done! Successfully processed {len(rows)} rows. Browser remains open.\n\n"

        except Exception as e:
            err_str = str(e)
            yield f"data: ❌ Error: {err_str}\n\n"
        finally:
            # Removed driver.quit() as requested to keep browser open
            pass

    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/extractFiles', methods=['GET', 'POST'])
def extract_files():
    from flask import Response, stream_with_context  # type: ignore[import-untyped]

    def generate():
        # --- load workbook ---
        if not os.path.exists(EXCEL_FILE):
            yield f"data: ❌ Excel file '{EXCEL_FILE}' not found.\n\n"
            return
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        except Exception as e:
            yield f"data: ❌ Failed to load workbook: {e}\n\n"
            return
        if 'Payment' not in wb.sheetnames:
            yield f"data: ❌ Sheet 'Payment' not found.\n\n"
            return

        sheet = wb['Payment']
        url_rows = []
        for r in range(2, sheet.max_row + 1):
            cell_val = sheet.cell(row=r, column=10).value
            if cell_val:
                url = str(cell_val).strip()
                if url.startswith('http'):
                    url_rows.append((r, url))

        if not url_rows:
            yield "data: ❌ No URLs found in column J of the Payment sheet.\n\n"
            return

        yield f"data: Found {len(url_rows)} links to process. Launching Chrome...\n\n"

        # --- launch Selenium ---
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from webdriver_manager.chrome import ChromeDriverManager
            import time

            opts = Options()
            opts.add_argument("--start-maximized")
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
        except Exception as e:
            yield f"data: ❌ Failed to launch Chrome: {e}\n\n"
            return

        try:
            # Land on Index.jsp — user logs in here
            driver.get("https://www.eprocure.gov.bd/Index.jsp")
            yield "data: ⏳ Waiting for you to log in (up to 3 minutes)...\n\n"

            # Detect login: URL leaves Index.jsp
            deadline = time.time() + 180
            logged_in = False
            while time.time() < deadline:
                cur = driver.current_url
                if 'Index.jsp' not in cur and 'Login' not in cur and 'eprocure.gov.bd' in cur:
                    logged_in = True
                    break
                time.sleep(1)

            if not logged_in:
                yield "data: ❌ Login timed out (3 minutes). Please restart and log in faster.\n\n"
                return

            yield "data: ✅ Login detected! Starting extraction...\n\n"
            time.sleep(1)

            updated: int = 0
            for idx, (r, url) in enumerate(url_rows, 1):
                yield f"data: [{idx}/{len(url_rows)}] Row {r}: Visiting link via simulated click...\n\n"
                try:
                    # Using simulated click to preserve Referer/Session
                    driver.execute_script(f"var a=document.createElement('a');a.href='{url}';a.style.display='none';document.body.appendChild(a);a.click();")
                    time.sleep(2)
                    # Use a shorter wait to check for session expiry
                    try:
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.CLASS_NAME, 'tableList_1'))
                        )
                    except:
                        # If table not found, maybe session expired? Try one more time to visit url
                        driver.get(url)
                        time.sleep(1)
                        if 'tableList_1' not in driver.page_source:
                            yield f"data: ⚠️  Row {r}: Table not found. If this persists, your session may have expired. Link unchanged.\n\n"
                            continue

                    soup = BeautifulSoup(driver.page_source, 'html.parser')
                    table = soup.find('table', class_='tableList_1')
                    if not table:
                        yield f"data: ⚠️  Row {r}: File table not found on page. Link unchanged.\n\n"
                        continue

                    data_trs = [tr for tr in table.find_all('tr') if tr.find('td')]
                    if not data_trs:
                        yield f"data: ⚠️  Row {r}: No data rows in table. Link unchanged.\n\n"
                        continue

                    tds = data_trs[-1].find_all('td')
                    if len(tds) >= 2:
                        file_name = tds[1].get_text(strip=True)
                        sheet.cell(row=r, column=10).value = file_name
                        try:
                            wb.save(EXCEL_FILE)
                            updated = updated + 1 # type: ignore[operator]
                            yield f"data: ✅ Row {r}: Saved → {file_name}\n\n"
                        except Exception as save_err:
                            sheet.cell(row=r, column=10).value = url
                            yield f"data: ⚠️  Row {r}: Got name but failed to save Excel: {save_err}\n\n"
                    else:
                        yield f"data: ⚠️  Row {r}: Not enough columns in last row. Link unchanged.\n\n"

                except Exception as inner_e:
                    err_msg = str(inner_e)
                    yield f"data: ⚠️  Row {r}: Exception — {err_msg[:120]}. Link unchanged.\n\n"

            yield f"data: \n🎉 Done! Successfully replaced {updated} / {len(url_rows)} links.\n\n"

        except Exception as e:
            yield f"data: ❌ Error: {str(e)}\n\n"
        finally:
            # Persistent browser as requested
            pass

    return Response(stream_with_context(generate()), mimetype='text/event-stream')

def open_browser():
    webbrowser.open_new('http://127.0.0.1:5000/')

if __name__ == '__main__':
    print("Starting Local Server...")
    print("A browser window will open shortly at http://127.0.0.1:5000/")
    Timer(1.5, open_browser).start()
    app.run(port=5000, debug=False)
